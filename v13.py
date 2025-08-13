import streamlit as st
import pandas as pd
import queue

# === Helper to keep DataFrames Arrow/Streamlit friendly ===
def _make_arrow_safe(df: pd.DataFrame) -> pd.DataFrame:
    try:
        import pandas as pd  # ensure pd is in scope
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_timedelta64_dtype(df[col]):
                # Use readable strings (or switch to minutes by uncommenting next line)
                df[col] = df[col].astype("string")
                # df[col] = (df[col].dt.total_seconds() / 60).round(2)
            elif df[col].dtype == "object":
                # Mixed objects: if any Timedelta present, cast all to string
                if any(isinstance(x, pd.Timedelta) for x in df[col].dropna()):
                    df[col] = df[col].astype("string")
        return df
    except Exception:
        return df

from datetime import datetime, timedelta, time
import matplotlib.pyplot as plt
import seaborn as sns
import requests
import io
from fpdf import FPDF
from PIL import Image
import sqlite3
import bcrypt
import os
from dotenv import load_dotenv
import base64
from io import BytesIO, StringIO
import json
import urllib3
from cryptography.fernet import Fernet, InvalidToken
from textwrap import dedent as _dedent

# New imports for Split by MZ and V3 functionality
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from tzlocal import get_localzone
from collections import defaultdict
import gc

# Import the deploy module to access its global variables and functions
import deploy
from deploy import onboard_agent

# Imports for Global Anomaly Detection
import openpyxl
from openpyxl.styles import Alignment

urllib3.disable_warnings()

DB_FILE = "users.db"

load_dotenv()

# Get the encryption key from environment variables
ENCRYPTION_KEY = os.getenv("ENCRYPTION_KEY")
if ENCRYPTION_KEY is None:
    st.error("Encryption key not found in .env file. Please generate one and add it.")
    st.stop()

cipher_suite = Fernet(ENCRYPTION_KEY.encode())

# Helper functions for encryption/decryption
def encrypt_token(token):
    if token:
        return cipher_suite.encrypt(token.encode()).decode()
    return None

def decrypt_token(encrypted_token):
    if encrypted_token:
        try:
            return cipher_suite.decrypt(encrypted_token.encode()).decode()
        except InvalidToken:
            st.error("Invalid encryption token. Could not decrypt API token.")
            return None
    return None

# ===== DATABASE FUNCTIONS =====
def create_connection():
    conn = sqlite3.connect(DB_FILE)
    return conn

def create_tables():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            hashed_password TEXT,
            role TEXT,
            mgmt_zones TEXT,
            dynatrace_url TEXT DEFAULT NULL,
            access_expiry TEXT DEFAULT NULL,
            encrypted_dynatrace_api_token TEXT DEFAULT NULL,
            access_authority BOOLEAN DEFAULT TRUE
        )''')
    try:
        c.execute('ALTER TABLE users ADD COLUMN dynatrace_url TEXT')
    except:
        pass
    try:
        c.execute('ALTER TABLE users ADD COLUMN access_expiry TEXT')
    except:
        pass
    try:
        c.execute("PRAGMA table_info(users)")
        columns = [col[1] for col in c.fetchall()]
        if "dynatrace_api_token" in columns and "encrypted_dynatrace_api_token" not in columns:
            c.execute('ALTER TABLE users RENAME COLUMN dynatrace_api_token TO encrypted_dynatrace_api_token')
        elif "encrypted_dynatrace_api_token" not in columns:
            c.execute('ALTER TABLE users ADD COLUMN encrypted_dynatrace_api_token TEXT')
    except Exception as e:
        print(f"Error adding/renaming column: {e}")
        pass
    try:
        c.execute('ALTER TABLE users ADD COLUMN access_authority BOOLEAN DEFAULT TRUE')
    except sqlite3.OperationalError:
        pass
    except Exception as e:
        print(f"Error adding access_authority column: {e}")
        pass
    conn.commit()
    conn.close()

def add_user(username, password, role="user", mgmt_zones=None, dynatrace_url=None, access_expiry=None, dynatrace_api_token=None):
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    zones = ",".join(mgmt_zones) if mgmt_zones else ""
    expiry_str = access_expiry.strftime("%Y-%m-%d") if access_expiry else None
    encrypted_token = encrypt_token(dynatrace_api_token)
    conn = create_connection()
    c = conn.cursor()
    try:
        c.execute(
            "INSERT INTO users (username, hashed_password, role, mgmt_zones, dynatrace_url, access_expiry, encrypted_dynatrace_api_token, access_authority) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (username, hashed, role, zones, dynatrace_url or "", expiry_str, encrypted_token or "", True))
        conn.commit()
    except sqlite3.IntegrityError:
        st.error("User already exists.")
        return False
    finally:
        conn.close()
    return True

def authenticate_user(username, password):
    if check_admin_credentials(username, password):
        return "admin", [], None, None
    conn = create_connection()
    c = conn.cursor()
    c.execute("SELECT hashed_password, role, mgmt_zones, dynatrace_url, access_expiry, encrypted_dynatrace_api_token, access_authority FROM users WHERE username = ?", (username,))
    user = c.fetchone()
    conn.close()
    if user:
        hashed, role, zones, dynatrace_url, expiry_date, encrypted_dynatrace_api_token, access_authority = user
        if bcrypt.checkpw(password.encode(), hashed.encode()):
            if not access_authority:
                st.error("Your access has been deactivated. Please contact the administrator.")
                return None, None, None, None
            if expiry_date:
                try:
                    expiry_dt = datetime.strptime(expiry_date, "%Y-%m-%d")
                    if datetime.now().date() > expiry_dt.date():
                        st.error("Your access has expired. Please contact the administrator.")
                        conn = create_connection()
                        c = conn.cursor()
                        c.execute("UPDATE users SET access_authority = ? WHERE username = ?", (False, username))
                        conn.commit()
                        conn.close()
                        return None, None, None, None
                except Exception as e:
                    st.warning("Invalid expiry date format.")
            zones_list = [z.strip() for z in (zones or '').split(",") if z.strip()]
            decrypted_api_token = decrypt_token(encrypted_dynatrace_api_token)
            return role, zones_list, dynatrace_url, decrypted_api_token
    return None, None, None, None

def get_all_users():
    conn = create_connection()
    c = conn.cursor()
    c.execute("SELECT id, username, role, mgmt_zones, dynatrace_url, access_expiry, access_authority FROM users")
    users = c.fetchall()
    conn.close()
    return users

def delete_user(username):
    conn = create_connection()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit()
    conn.close()

def update_user_expiry(username, new_expiry_date):
    conn = create_connection()
    c = conn.cursor()
    expiry_str = new_expiry_date.strftime("%Y-%m-%d")
    try:
        c.execute("UPDATE users SET access_expiry = ?, access_authority = ? WHERE username = ?", (expiry_str, True, username))
        conn.commit()
        return True
    except Exception as e:
        st.error(f"Error updating user expiry: {e}")
        return False
    finally:
        conn.close()

def check_admin_credentials(username, password):
    admin_user = os.getenv("ADMIN_USERNAME")
    admin_pass = os.getenv("ADMIN_PASSWORD")
    return username == admin_user and password == admin_pass

create_tables()

def check_session_timeout():
    if 'auth_user' in st.session_state:
        login_time = st.session_state.get('login_time')
        if login_time:
            if datetime.now() - login_time > timedelta(hours=24):
                st.warning("⏰ Session expired. Please login again.")
                logout()
                return False
        else:
            st.session_state['login_time'] = datetime.now()
    return True

# ===== PAGE CONFIG & BASE CSS with Animations =====
st.set_page_config(layout="wide", page_title="Dynatrace Problem Dashboard", initial_sidebar_state="expanded")


st.markdown("""
<style>
:root { --accent:#2196f3; }
.stButton>button{
  border-radius:10px;color:#fff;background:var(--accent);
  transition:transform .18s ease, box-shadow .18s ease; 
  box-shadow:0 6px 14px rgba(0,0,0,.25);
}
.stButton>button:hover{ transform:translateY(-2px) scale(1.01); box-shadow:0 10px 22px rgba(0,0,0,.33); }
.title-wrap{ margin-top:.5rem }
.title-pro{ font-size:42px; margin:0; line-height:1.1; }
.title-shimmer{
  background: linear-gradient(90deg, #fff 0%, #9ecbff 20%, #fff 40%);
  background-size:200% 100%;
  -webkit-background-clip:text; background-clip:text; color:transparent;
  animation: shimmer 2.2s ease-in-out infinite;
}
.title-underline{
  width:0;height:3px;background:var(--accent); border-radius:999px; margin-top:8px;
  animation: sweep 1s ease .4s forwards;
}
@keyframes shimmer{ 0%{background-position:200% 0} 100%{background-position:-200% 0} }
@keyframes sweep{ to{ width: 280px } }
.kpi{
  background:linear-gradient(180deg,#2a2f3a 0%,#20242d 100%);
  padding:16px;border-radius:14px;box-shadow:0 8px 24px rgba(0,0,0,.25);
  border:1px solid rgba(255,255,255,.06);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.kpi:hover{ transform:translateY(-2px); box-shadow:0 12px 32px rgba(0,0,0,.32); border-color: rgba(33,150,243,.35) }
.kpi-title{ font-size:13px; opacity:.8; margin-bottom:6px }
.kpi-value{ font-size:28px; font-weight:700 }
.kpi-trend{ font-size:12px; opacity:.85 }
.skel{
  border-radius:10px; height:18px; background:linear-gradient(90deg,#313645 25%,#3a4154 37%,#313645 63%);
  background-size:400% 100%; animation: sk 1.1s ease-in-out infinite;
}
@keyframes sk{ 0%{background-position:100% 0} 100%{background-position:-100% 0} }
.sticky-bar{ position:sticky; top:0; z-index: 999; backdrop-filter: blur(6px);
  background: rgba(24,26,33,.66); border-bottom:1px solid rgba(255,255,255,.08);
  padding:8px 10px; border-radius: 12px; margin: 8px 0 12px 0; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
button[title="View fullscreen"],
button[aria-label="View fullscreen"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
    <style>
    /* Base Styles */
    .stButton>button {
        border-radius: 8px;
        color: white;
        background: #2196f3;
        transition: all 0.2s ease-in-out;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 10px rgba(0, 0, 0, 0.3);
    }
    .stSelectbox>div>div {
        background: #21232b;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }
    .stTextInput>div>input {
        background: #1c1f26;
        color: white;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        transition: box-shadow 0.2s ease;
    }
    .stTextInput>div>input:focus {
        box-shadow: 0 4px 8px rgba(33, 150, 243, 0.3);
    }
    .stDataFrame, .stTable {
        background: #252834;
        color: white;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
        border-radius: 8px;
        overflow: hidden;
    }
    footer { visibility: hidden; }

    /* Custom Animations for Logo and Title */
    @keyframes moveAndFadeIn {
        0% { transform: translateX(-100px); opacity: 0; }
        100% { transform: translateX(0); opacity: 1; }
    }
    @keyframes pulseFadeIn {
      0% { transform: scale(0.95); opacity: 0; }
      50% { transform: scale(1.05); opacity: 1; }
      100% { transform: scale(1); opacity: 1; }
    }
    
    .logo-animation {
        animation: moveAndFadeIn 1.5s ease-out;
    }
    .title-animation {
        animation: pulseFadeIn 1s ease-out 0.5s both;
    }
    
    /* Login Page Animation */
    @keyframes slideInUp {
      0% { transform: translateY(20px); opacity: 0; }
      100% { transform: translateY(0); opacity: 1; }
    }
    .slide-in-up {
      animation: slideInUp 0.8s ease-out;
    }
    
    </style>
""", unsafe_allow_html=True)

# ===== AUTHENTICATION (SIGNUP / LOGIN / LOGOUT) with Animations =====
def get_base64_image(img):
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()

def login():
    logo_file = "apmolens_logo.png"
    try:
        logo = Image.open(logo_file)
        logo_base64 = get_base64_image(logo)
        st.markdown(
            f"""
            <div class="slide-in-up" style="display: flex; flex-direction: column; align-items: center; justify-content: center; margin-bottom: 1.5rem;">
                <img src="data:image/png;base64,{logo_base64}" width="320" alt="APM Lens Logo"/>
            </div>
            """,
            unsafe_allow_html=True,
        )
    except (FileNotFoundError, IOError, Exception) as e:
        st.markdown(
            """
            <div class="slide-in-up" style="display: flex; flex-direction: column; align-items: center; justify-content: center; margin-bottom: 1.5rem;">
            </div>
            """,
            unsafe_allow_html=True,
        )
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown(
            """
            <h2 style="color: #fff; text-align: center; margin-bottom: 0.5rem; font-weight:600;">Login</h2>
            """,
            unsafe_allow_html=True,
        )
        with st.container():
            st.markdown(
                """
                <div class="slide-in-up">
                """,
                unsafe_allow_html=True
            )
            with st.form("login_form"):
                username = st.text_input(
                    "Username",
                    key="login_username",
                    label_visibility="visible",
                    placeholder="Enter your username"
                )
                password = st.text_input(
                    "Password",
                    type="password",
                    key="login_password",
                    label_visibility="visible",
                    placeholder="Enter your password"
                )
                login_button = st.form_submit_button("Login", use_container_width=True)
            if login_button:
                if not username or not password:
                    st.error("Please fill in both username and password.", icon="⚠️")
                else:
                    role, zones, dynatrace_url, dynatrace_api_token = authenticate_user(username, password)
                    if role:
                        st.session_state['auth_user'] = username
                        st.session_state['auth_role'] = role
                        st.session_state['auth_zones'] = zones
                        st.session_state['auth_dt_url'] = dynatrace_url
                        st.session_state['auth_dt_api_token'] = dynatrace_api_token
                        st.session_state['login_time'] = datetime.now()
                        st.success(f"Welcome, {username}!", icon="✅")
                        st.rerun()
                    else:
                        st.error("Invalid credentials. Please try again.", icon="⚠️")
    st.markdown("""
        <style>
            div[data-testid="stForm"] {border: 1px solid rgba(49, 51, 63, 0.3); border-radius: 0.75rem; padding: 1.5rem; max-width: 350px; margin: 0 auto; background: rgba(28, 31, 38, 0.6); backdrop-filter: blur(10px); box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);}
            .stTextInput > div > input {font-size: 14px !important; height: 35px !important; padding: 8px 12px !important; border-radius: 6px !important; border: 1px solid rgba(255, 255, 255, 0.2) !important; background: rgba(28, 31, 38, 0.8) !important; color: white !important; transition: border-color 0.3s ease !important;}
            .stTextInput > div > input:focus {border-color: #2196f3 !important; box-shadow: 0 0 0 2px rgba(33, 150, 243, 0.2) !important;}
            .stFormSubmitButton > button {background: rgba(70, 70, 70, 0.8) !important; border: 1px solid rgba(255, 255, 255, 0.2) !important; border-radius: 6px !important; height: 40px !important; font-weight: 600 !important; color: white !important; transition: all 0.3s ease !important;}
            .stFormSubmitButton > button:hover {background: rgba(90, 90, 90, 0.9) !important; border-color: rgba(255, 255, 255, 0.4) !important; transform: translateY(-1px) !important; box-shadow: 0 4px 8px rgba(0, 0, 0, 0.3) !important;}
            div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] {gap: 0.75rem !important;}
            .stTextInput > label {color: rgba(255, 255, 255, 0.9) !important; font-weight: 500 !important;}
        </style>
    """, unsafe_allow_html=True)
    st.stop()

def logout():
    st.session_state.clear()
    st.session_state['rerun_after_logout'] = True

if "auth_user" not in st.session_state:
    login()

if not check_session_timeout():
    st.stop()

# ===== MANAGEMENT ZONES HELPERS =====
def fetch_management_zones(env_id, api_token):
    url = f"{env_id.rstrip('/')}/api/v2/settings/objects"
    headers = {'Authorization': f'Api-Token {api_token}', 'Content-Type': 'application/json'}
    params = {'schemaIds': 'builtin:management-zones', 'pageSize': 10}
    try:
        response = requests.get(url, headers=headers, params=params, verify=False, timeout=30)
        if response.status_code == 200:
            items = response.json().get('items', [])
            zones = [zone['value']['name'] for zone in items]
            return zones
        elif response.status_code == 401:
            st.error("❌ Authentication failed. Please check your API token.")
        elif response.status_code == 403:
            st.error("❌ Access forbidden. API token may not have required permissions.")
        elif response.status_code == 404:
            st.error("❌ Environment not found. Please check your Dynatrace URL.")
        else:
            st.error(f"❌ API Error {response.status_code}: {response.text}")
    except requests.exceptions.Timeout:
        st.error("❌ Request timed out. Please try again.")
    except requests.exceptions.ConnectionError:
        st.error("❌ Connection error. Please check your network and Dynatrace URL.")
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Request failed: {str(e)}")
    except Exception as e:
        st.error(f"❌ Unexpected error: {str(e)}")
    return []

def import_mz_from_csv(uploaded_file):
    try:
        if uploaded_file is not None:
            df = pd.read_csv(uploaded_file)
            maj_cols = [c.lower() for c in df.columns]
            if 'name' in maj_cols:
                name_col = df.columns[maj_cols.index('name')]
                return sorted(df[name_col].dropna().astype(str).unique())
            else:
                return sorted(df[df.columns[0]].dropna().astype(str).unique())
    except Exception as e:
        st.error(f"CSV read error: {e}")
    return []

# ===== MAINTENANCE WINDOW FUNCTIONS =====
def fetch_management_zones_mw(base_url, token):
    url = f"{base_url.rstrip('/')}/api/v2/settings/objects"
    headers = {'Authorization': f'Api-Token {token}', 'Content-Type': 'application/json'}
    params = {'schemaIds': 'builtin:management-zones', 'pageSize': 10}
    try:
        resp = requests.get(url, headers=headers, params=params, verify=False)
        if resp.status_code == 200:
            zones = {item['value']['name']: item['objectId'] for item in resp.json().get('items', [])}
            return zones
        else:
            st.error(f"Failed to fetch zones: {resp.text}")
            return {}
    except Exception as e:
        st.error(f"Error: {e}")
        return {}

def fetch_hosts_mw(base_url, token, selected_zone_names, all_zones):
    zone_ids = [all_zones[name] for name in selected_zone_names]
    entity_selector = 'type(HOST)'
    if zone_ids:
        for zone_id in zone_ids:
            entity_selector += f",mzId({zone_id})"
    url = f"{base_url.rstrip('/')}/api/v2/entities"
    headers = {'Authorization': f'Api-Token {token}', 'Content-Type': 'application/json'}
    params = {'entitySelector': entity_selector, 'pageSize': 10}
    try:
        resp = requests.get(url, headers=headers, params=params, verify=False)
        if resp.status_code == 200:
            hosts = {item['displayName']: item['entityId'] for item in resp.json().get('entities', [])}
            return hosts
        else:
            st.error(f"Failed to fetch hosts: {resp.text}")
            return {}
    except Exception as e:
        st.error(f"Error: {e}")
        return {}

def create_maintenance_window_api(base_url, token, name, desc, start_time, end_time,
                               selected_host_names, all_hosts,
                               selected_zone_names, all_zones):
    entity_ids = [all_hosts[name] for name in selected_host_names] if selected_host_names else []
    zone_ids = [all_zones[name] for name in selected_zone_names] if selected_zone_names else []
    try:
        start_timestamp = int(datetime.strptime(start_time, "%Y-%m-%d %H:%M").timestamp() * 1000)
        end_timestamp = int(datetime.strptime(end_time, "%Y-%m-%d %H:%M").timestamp() * 1000)
    except ValueError:
        st.error("Invalid date format. Use YYYY-MM-DD HH:MM")
        return False
    scope_matches = []
    if entity_ids:
        scope_matches.append({"type": "entities", "entities": entity_ids})
    if zone_ids:
        scope_matches.extend([{"type": "mzId", "mzId": zid} for zid in zone_ids])
    payload = {
        "type": "PLANNED",
        "name": name,
        "description": desc,
        "suppression": {"problems": True, "alerts": True},
        "scope": {"matches": scope_matches},
        "schedule": {"type": "ONCE", "start": start_timestamp, "end": end_timestamp, "zoneId": "Asia/Kolkata"}
    }
    url = f"{base_url.rstrip('/')}/api/v2/maintenanceWindows"
    headers = {'Authorization': f'Api-Token {token}', 'Content-Type': 'application/json'}
    try:
        resp = requests.post(url, headers=headers, data=json.dumps(payload), verify=False)
        if resp.status_code in [200, 201]:
            return True
        else:
            st.error(f"Failed to create maintenance window: {resp.status_code} - {resp.text}")
            return False
    except Exception as e:
        st.error(f"Error: {e}")
        return False

def maintenance_window_ui(env_id, api_token):
    st.header("Create Dynatrace Maintenance Window")
    base_url = st.text_input("Dynatrace Base URL", value=env_id, key="mw_base_url_input", disabled=True)
    token = st.text_input("API Token", value="********", type="password", key="mw_api_token_input", disabled=True)
    if st.button("Fetch Management Zones for MW", key="fetch_mz_mw_btn"):
        if base_url and api_token:
            zones = fetch_management_zones_mw(base_url, api_token)
            st.session_state.mw_zones = zones
            if zones:
                st.success(f"Fetched {len(zones)} management zones.")
            else:
                st.warning("No management zones found or an error occurred.")
        else:
            st.error("Please provide Dynatrace Base URL and API Token.")
    selected_zones = []
    if "mw_zones" in st.session_state and st.session_state.mw_zones:
        selected_zones = st.multiselect("Select Management Zones", list(st.session_state.mw_zones.keys()), key="mw_selected_zones")
        if selected_zones and st.button("Fetch Hosts in Selected Zones", key="fetch_hosts_mw_btn"):
            if base_url and api_token:
                hosts = fetch_hosts_mw(base_url, api_token, selected_zones, st.session_state.mw_zones)
                st.session_state.mw_hosts = hosts
                if hosts:
                    st.success(f"Fetched {len(hosts)} hosts.")
                else:
                    st.warning("No hosts found in selected zones or an error occurred.")
            else:
                st.error("Please provide Dynatrace Base URL and API Token.")
    elif "mw_zones" in st.session_state and not st.session_state.mw_zones:
        st.info("No hosts available. Fetch them first.")
    selected_hosts = []
    if "mw_hosts" in st.session_state and st.session_state.mw_hosts:
        selected_hosts = st.multiselect("Select Hosts", list(st.session_state.mw_hosts.keys()), key="mw_selected_hosts")
    elif "mw_hosts" in st.session_state and not st.session_state.mw_hosts:
        st.info("No hosts available. Fetch them first.")
    st.markdown("---")
    st.subheader("Maintenance Window Details")
    name = st.text_input("Maintenance Window Name", key="mw_name")
    desc = st.text_input("Description", key="mw_desc")
    col_start, col_end = st.columns(2)
    with col_start:
        start_date = st.date_input("Start Date", value=datetime.today(), key="mw_start_date")
        start_time_val = datetime.now().time()
        start_time_str = st.time_input("Start Time", value=start_time_val, key="mw_start_time").strftime("%H:%M")
    with col_end:
        end_date = st.date_input("End Date", value=datetime.today() + timedelta(hours=1), key="mw_end_date")
        end_time_val = (datetime.now() + timedelta(hours=1)).time()
        end_time_str = st.time_input("End Time", value=end_time_val, key="mw_end_time").strftime("%H:%M")
    full_start_time = f"{start_date.strftime('%Y-%m-%d')} {start_time_str}"
    full_end_time = f"{end_date.strftime('%Y-%m-%d')} {end_time_str}"
    st.info(f"Maintenance Window will be from: **{full_start_time}** to **{full_end_time}**")
    if st.button("Create Maintenance Window", key="create_mw_btn"):
        if not name or not desc:
            st.error("Please fill in Maintenance Window Name and Description.")
        elif not base_url or not api_token:
            st.error("Please provide Dynatrace Base URL and API Token.")
        elif not selected_hosts and not selected_zones:
            st.error("Please select at least one Host or Management Zone.")
        else:
            success = create_maintenance_window_api(
                base_url, api_token, name, desc, full_start_time, full_end_time,
                selected_hosts, st.session_state.mw_hosts if "mw_hosts" in st.session_state else {},
                selected_zones, st.session_state.mw_zones if "mw_zones" in st.session_state else {}
            )
            if success:
                st.success("Maintenance window created successfully.")
                st.session_state.mw_name = ""
                st.session_state.mw_desc = ""
                st.session_state.mw_selected_zones = []
                st.session_state.mw_selected_hosts = []
                st.session_state.mw_hosts = {}
                st.session_state['rerun_flag'] = True
            else:
                st.error("Failed to create maintenance window. Check logs for details.")

def batch_deploy_one_agent(df, num_threads=10):
    """
    Orchestrates the deployment of OneAgent from a DataFrame of server configurations
    using multithreading. Results are collected and returned.
    
    The function now reads a 'mode' column from the DataFrame to set the monitoring mode.
    If 'mode' is not present, it defaults to 'full'.
    """
    task_queue = queue.Queue()
    results_queue = queue.Queue()  # To collect results from threads

    # Populate the queue with deployment tasks
    for index, row in df.iterrows():
        # Ensure row['password'] is not empty before putting it on the queue
        # For SSH key auth, you would need to handle that in the CSV too.
        # This example assumes password auth for batch
        if 'password' not in row or pd.isna(row['password']):
            results_queue.put({"host": row['host'], "status": "error", "message": "Password column is missing or empty."})
            continue
        task_queue.put(row)

    def worker():
        while True:
            try:
                row = task_queue.get(timeout=1)
                host_ip = row['host']
                
                try:
                    # Pass tenant_url as is from CSV, deploy.py will normalize it
                    # Also, pass the mode from the CSV, defaulting to 'full'
                    deploy.onboard_agent(
                        dt_tenant=row['tenant_url'],
                        dt_api_token=row['api_token'],
                        host_ip=host_ip,
                        username=row['username'],
                        password=row['password'],
                        host_group=row.get('host_group', 'Default'),
                        mode=row.get('mode', 'full') # Correctly retrieves mode, defaults to 'full'
                    )
                    results_queue.put({"host": host_ip, "status": "success", "message": "completed successfully."})
                except Exception as e:
                    results_queue.put({"host": host_ip, "status": "error", "message": str(e)})
                finally:
                    task_queue.task_done()

            except queue.Empty:
                break
            except Exception as e:
                results_queue.put({"host": "UNKNOWN", "status": "error", "message": f"Worker error: {str(e)}"})
                task_queue.task_done()

    threads = []
    for i in range(num_threads):
        thread = threading.Thread(target=worker, name=f"Worker-{i+1}")
        thread.daemon = True
        threads.append(thread)
        thread.start()

    task_queue.join()

    all_results = []
    while not results_queue.empty():
        all_results.append(results_queue.get())
    
    return all_results

# -----------------------------------------------------------------------------

def render_oneagent_deployment_tab():
    st.header("Dynatrace OneAgent Deployment")
    st.markdown("Automate the deployment of the Dynatrace OneAgent to a target server.")

    deployment_method = st.radio(
        "Select Deployment Method",
        ["Manual Input", "Upload CSV for Batch Deployment"]
    )
    
    if deployment_method == "Manual Input":
        with st.form("oneagent_deployment_form"):
            st.subheader("Server Credentials")
            server_ip = st.text_input("Server Hostname or IP", key="deploy_server_ip")
            username = st.text_input("SSH Username", key="deploy_username")
            
            auth_method = st.radio("Authentication Method", ["Password", "SSH Key"], key="auth_method")

            password = None
            private_key = None
            if auth_method == "Password":
                password = st.text_input("SSH Password", type="password", key="deploy_password")
            else:
                private_key = st.text_area("SSH Private Key", height=200, key="deploy_private_key")

            st.subheader("Dynatrace Configuration")
            dynatrace_url = st.text_input(
                "Dynatrace Environment URL",
                value=st.session_state.get('auth_dt_url', ''),
                key="deploy_dt_url"
            )
            api_token = st.text_input(
                "Dynatrace API Token",
                type="password",
                help="API Token with 'Installer download' permission.",
                key="deploy_api_token"
            )
            
            st.subheader("Deployment Options")
            col1, col2 = st.columns(2)
            with col1:
                os_type = st.selectbox("Target OS", options=["unix", "windows"], key="deploy_os_type")
            with col2:
                arch = st.selectbox("Architecture", options=["x86-64"], key="deploy_arch")

            col3, col4 = st.columns(2)
            with col3:
                host_group = st.text_input("Host Group (Optional)", key="deploy_host_group", help="Specify a host group to tag the deployed agent.")
            
            with col4:
                monitoring_mode = st.selectbox(
                    "Monitoring Mode",
                    options=["full", "infra"],
                    help="Select 'full' for full-stack monitoring or 'infra' for infrastructure-only.",
                    key="deploy_monitoring_mode"
                )

            submit_button = st.form_submit_button("Deploy OneAgent", use_container_width=True)

        if submit_button:
            if not server_ip or not username or not dynatrace_url or not api_token:
                st.error("Please fill in all required fields.")
            elif auth_method == "Password" and not password:
                st.error("Please enter a password.")
            elif auth_method == "SSH Key" and not private_key:
                st.error("Please enter an SSH private key.")
            else:
                try:
                    # Pass the correct auth method based on user selection
                    deploy.onboard_agent(
                        dt_tenant=dynatrace_url,
                        dt_api_token=api_token,
                        host_ip=server_ip,
                        username=username,
                        password=password,
                        # Pass private_key to onboard_agent if it's not None
                        private_key=private_key, 
                        host_group=host_group,
                        mode=monitoring_mode
                    )
                    st.success(f"Deployment to {server_ip} completed successfully.")
                except Exception as e:
                    st.error(f"Deployment to {server_ip} failed: {str(e)}")
    
    else: # Upload CSV for Batch Deployment
        st.info("""
        Please upload a CSV file with the following columns:
        `tenant_url`, `api_token`, `host`, `username`, `password`, `host_group` (optional), and `mode` (optional, 'full' or 'infra').
        If `mode` is not specified, it will default to `full` for each host.
        A template is provided for download.
        """)
        
        csv_template = pd.DataFrame({
            'tenant_url': ['https://your-env.live.dynatrace.com'],
            'api_token': ['your-api-token-with-installer-download-permission'],
            'host': ['10.0.0.1'],
            'username': ['user'],
            'password': ['password'],
            'host_group': ['web-servers'],
            'mode': ['full']
        })
        st.download_button(
            label="Download CSV Template",
            data=csv_template.to_csv(index=False),
            file_name="oneagent_deployment_template.csv",
            mime="text/csv"
        )

        uploaded_file = st.file_uploader("Upload Deployment CSV", type=["csv"])
        if uploaded_file:
            try:
                df = pd.read_csv(uploaded_file)
                st.write("### Uploaded Data")
                st.dataframe(df)

                if st.button("Start Batch Deployment from CSV"):
                    required_columns = ['tenant_url', 'api_token', 'host', 'username']
                    if not all(col in df.columns for col in required_columns):
                        st.error(f"CSV is missing required columns. Please ensure it contains: {', '.join(required_columns)}")
                    elif 'password' not in df.columns:
                        st.error("CSV is missing the required 'password' column for batch deployment.")
                    else:
                        with st.spinner("Starting batch deployment... This may take a while."):
                            deployment_results = batch_deploy_one_agent(df, num_threads=10)
                        
                        st.subheader("Deployment Results Summary")
                        for result in deployment_results:
                            if result["status"] == "success":
                                st.success(f"✅ Host {result['host']}: {result['message']}")
                            else:
                                st.error(f"❌ Host {result['host']}: {result['message']}")
                        
                        if not deployment_results:
                            st.info("No deployment tasks were processed.")
                        else:
                            st.success("Batch deployment process completed.")

            except Exception as e:
                st.error(f"Error reading CSV file: {e}")


# ===== ADMIN TABS =====
def render_add_user_tab():
    st.header("Add User")
    # Essentials
    new_user = st.text_input("User Name", key="anewuser_main")
    new_user_pw = st.text_input("Password", type="password", key="anewpw_main")
    new_role = st.selectbox("Role", options=["user", "admin"], key="aroleuser_main")
    expiry_date = st.date_input("Access Expiry Date (User will be deactivated after this date)", key="expiry_date_main")
    # Advanced (collapsed by default)
    with st.expander("Advanced: Management Zones (Admin)", expanded=False):
        st.markdown("_Zones can be loaded from CSV here. Environment URL and API Token are managed on the main dashboard._")
        uploaded_csv = st.file_uploader("Upload CSV with a 'Management Zone' column", type=["csv"], key="mz_csv_main")
        dynamic_zone_options = []
        if uploaded_csv is not None:
            try:
                df = pd.read_csv(uploaded_csv)
                dynamic_zone_options = sorted(df['Management Zone'].dropna().astype(str).unique().tolist()) if 'Management Zone' in df.columns else []
                st.success(f"Loaded {len(dynamic_zone_options)} zones from CSV.")
            except Exception as e:
                st.error(f"CSV read error: {e}")
        selected_zones = st.multiselect("Assign Management Zones", options=dynamic_zone_options, key="assign_mz_main")
    # Save action
    if st.button("Create User"):
        zones_json = json.dumps(st.session_state.get('assign_mz_main', []))
        add_user(new_user, new_user_pw, new_role, zones_json, st.session_state.get('dt_url_main', ''), expiry_date)
        st.success(f"User {new_user} created.")

def admin_panel_tab():
    st.header("Manage Users")
    if 'manage_section' not in st.session_state:
        st.session_state['manage_section'] = 'Show Users'
    manage_section = st.radio('Section', ['Show Users', 'Renew User Access', 'Delete User'], index=['Show Users','Renew User Access','Delete User'].index(st.session_state['manage_section']), horizontal=True)
    st.session_state['manage_section'] = manage_section
    if manage_section == 'Show Users':
        st.subheader('User List')
        users = get_all_users()
        if users:
            user_data = []
            for uid, uname, role, zones, dt_url, expiry_date, access_authority in users:
                status = 'Active' if access_authority else 'Expired'
                user_data.append({'Username': uname, 'Role': role, 'Dynatrace URL': dt_url or 'N/A', 'Access Expiry': expiry_date or 'Never', 'Status': status})
            users_df = pd.DataFrame(user_data)
            st.dataframe(_make_arrow_safe(users_df), use_container_width=True)
        else:
            st.info('No users found in the database.')
    elif manage_section == 'Delete User':
        st.subheader('Delete User')
        all_users = [(uid, uname) for uid, uname, role, zones, dt_url, expiry_date, access_authority in get_all_users() if uname != st.session_state['auth_user']]
        uname_options = [uname for uid, uname in all_users]
        if uname_options:
            deluser = st.selectbox('Select user to delete', options=uname_options)
            if st.button('Delete Selected User') and deluser:
                delete_user(deluser)
                st.success(f'Deleted user {deluser}.')
        else:
            st.info('No deletable users found.')
    elif manage_section == 'Renew User Access':
        st.subheader('Renew / Update Access Expiry')
        users = [(uid, uname) for uid, uname, role, zones, dt_url, expiry_date, access_authority in get_all_users()]
        if users:
            uname_options = [uname for uid, uname in users]
            sel_uname = st.selectbox('Select user', options=uname_options)
            new_expiry = st.date_input('New expiry date')
            if st.button('Update Expiry'):
                update_user_expiry(sel_uname, new_expiry)
                st.success(f'Updated expiry for {sel_uname} to {new_expiry}.')
        else:
            st.info('No users found.')

def sanitize_data_for_excel(value):
    if value is None:
        return "N/A"
    str_value = str(value)
    str_value = ''.join(char for char in str_value if ord(char) >= 32 or char in '\t\n\r')
    str_value = re.sub(r'\s+', ' ', str_value).strip()
    if len(str_value) > 32000:
        str_value = str_value[:32000] + "... [truncated]"
    if str_value.startswith('='):
        str_value = "'" + str_value
    return str_value

def safe_hyperlink(ws, row, col, url, display_text):
    try:
        if url and len(str(url)) < 2000:
            clean_url = str(url).strip()
            if clean_url.startswith(('http://', 'https://')):
                ws.cell(row=row, column=col, value=display_text)
                ws.cell(row=row, column=col).hyperlink = clean_url
                ws.cell(row=row, column=col).font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=row, column=col, value=display_text)
        else:
            ws.cell(row=row, column=col, value=display_text)
    except Exception as e:
        print(f"Error adding hyperlink: {e}")
        ws.cell(row=row, column=col, value=display_text)

def sanitize_sheet_name(name):
    if not name:
        return "Sheet1"
    invalid_chars = ['\\', '/', '*', '?', '[', ']', ':']
    sanitized = str(name)
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    sanitized = sanitized.strip()
    if not sanitized:
        sanitized = "Sheet1"
    return sanitized

def fetch_problems_batch(base_url, headers, params, batch_size=500):
    all_problems = []
    def fetch_page(page_params):
        try:
            response = requests.get(base_url, params=page_params, headers=headers, verify=False)
            if response.status_code == 200:
                return response.json()
            else:
                print(f"Error fetching page: {response.status_code} - {response.text}")
                return None
        except Exception as e:
            print(f"Exception fetching page: {e}")
            return None
    try:
        response = requests.get(base_url, params=params, headers=headers, verify=False)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch data: {response.text}")
        data = response.json()
        all_problems.extend(data.get("problems", []))
        next_page_key = data.get("nextPageKey")
        if not next_page_key:
            return all_problems
        page_keys = [next_page_key]
        while True:
            temp_params = {"nextPageKey": next_page_key}
            temp_response = requests.get(base_url, params=temp_params, headers=headers, verify=False)
            if temp_response.status_code == 200:
                temp_data = temp_response.json()
                next_page_key = temp_data.get("nextPageKey")
                if next_page_key:
                    page_keys.append(next_page_key)
                else:
                    break
            else:
                break
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_params = {
                executor.submit(fetch_page, {"nextPageKey": page_key}): page_key
                for page_key in page_keys
            }
            for future in as_completed(future_to_params):
                result = future.result()
                if result:
                    all_problems.extend(result.get("problems", []))
    except Exception as e:
        print(f"Error in fetch_problems_batch: {e}")
        raise
    return all_problems

def calculate_mttr_data(unique_problems, selected_zones, from_date, to_date):
    try:
        mttr_data = []
        current_time = datetime.now()
        if hasattr(from_date, 'date'):
            from_date_obj = from_date
        else:
            from_date_obj = datetime.combine(from_date, datetime.min.time())
        if hasattr(to_date, 'date'):
            to_date_obj = to_date
        else:
            to_date_obj = datetime.combine(to_date, datetime.min.time())
        last_7_days_start = to_date_obj - timedelta(days=7)
        for idx, mgmt_zone in enumerate(selected_zones, 1):
            zone_problems = []
            zone_problems_last_7_days = []
            for prob_id, prob_info in unique_problems.items():
                if mgmt_zone in prob_info['management_zones']:
                    if prob_info['end_time'] is not None:
                        duration_minutes = prob_info['duration_minutes']
                    else:
                        start_time = prob_info['start_time']
                        duration_minutes = (current_time - start_time).total_seconds() / 60
                    if duration_minutes is not None and duration_minutes > 0:
                        zone_problems.append(duration_minutes)
                        problem_start = prob_info['start_time']
                        if hasattr(problem_start, 'date'):
                            problem_date = problem_start.date()
                        else:
                            problem_date = problem_start
                        if hasattr(last_7_days_start, 'date'):
                            last_7_days_date = last_7_days_start.date()
                        else:
                            last_7_days_date = last_7_days_start
                        if problem_date >= last_7_days_date:
                            zone_problems_last_7_days.append(duration_minutes)
            mttr_full = sum(zone_problems) / len(zone_problems) if zone_problems else 0
            mttr_last_7_days = sum(zone_problems_last_7_days) / len(zone_problems_last_7_days) if zone_problems_last_7_days else 0
            mttr_data.append({
                'Sr No': idx,
                'Management Zone': sanitize_data_for_excel(mgmt_zone),
                'MTTR (Minutes)': round(mttr_full, 2),
                'Last 7 Days MTTR (Minutes)': round(mttr_last_7_days, 2)
            })
        return mttr_data
    except Exception as e:
        print(f"Error in calculate_mttr_data: {e}")
        raise

def process_data_and_create_excel(environment_id, token, selected_zones, from_date, from_time, to_date, to_time):
    try:
        def convert_to_timestamp(date, time_obj):
            datetime_str = f"{date.strftime('%Y-%m-%d')} {time_obj.strftime('%H:%M')}"
            local_dt = pd.to_datetime(datetime_str).tz_localize(get_localzone())
            utc_dt = local_dt.tz_convert('UTC')
            return int(utc_dt.timestamp() * 1000)
        from_time_timestamp = convert_to_timestamp(from_date, from_time)
        to_time_timestamp = convert_to_timestamp(to_date, to_time)
        management_zones = ", ".join([f'"{zone}"' for zone in selected_zones])
        problem_selector = f'managementZones({management_zones})'
        base_url = f'{environment_id}/api/v2/problems'
        headers = {"Authorization": f"Api-Token {token}"}
        params = {
            "from": from_time_timestamp,
            "to": to_time_timestamp,
            "problemSelector": problem_selector,
            "pageSize": 10
        }
        local_timezone = get_localzone()
        current_time = datetime.now()
        st.info("Fetching problems. This may take a moment...")
        all_problems = fetch_problems_batch(base_url, headers, params)
        st.success(f"Fetched {len(all_problems)} problems.")
        unique_problems = {}
        for problem in all_problems:
            try:
                problem_id = problem.get("problemId")
                if not problem_id:
                    continue
                start_time = pd.to_datetime(problem.get("startTime"), unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                end_time = pd.to_datetime(problem.get("endTime"), unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None) if problem.get("endTime") != -1 else None
                if end_time is not None:
                    duration_minutes = (end_time - start_time).total_seconds() / 60
                else:
                    duration_minutes = (current_time - start_time).total_seconds() / 60
                problem_mgmt_zones = [zone.get("name", "N/A") for zone in problem.get("managementZones", [])]
                if problem_id not in unique_problems:
                    unique_problems[problem_id] = {
                        'start_time': start_time,
                        'end_time': end_time,
                        'duration_minutes': duration_minutes,
                        'management_zones': problem_mgmt_zones,
                        'problem_data': problem
                    }
            except Exception as e:
                print(f"Error processing problem {problem.get('problemId', 'unknown')}: {e}")
                continue
        mttr_data = calculate_mttr_data(unique_problems, selected_zones, from_date, to_date)
        zone_problems = {zone: [] for zone in selected_zones}
        zone_processed_problems = {zone: set() for zone in selected_zones}
        for problem in all_problems:
            try:
                problem_id = problem.get("problemId")
                if not problem_id:
                    continue
                start_time = pd.to_datetime(problem.get("startTime"), unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                end_time = pd.to_datetime(problem.get("endTime"), unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None) if problem.get("endTime") != -1 else None
                problem_url = f"{environment_id}/#problems/problemdetails;pid={problem.get('problemId')}"
                if end_time is not None:
                    active_time = str(end_time - start_time)
                    end_datetime_formatted = end_time.strftime('%m/%d/%Y %H:%M:%S')
                else:
                    active_time = f"Ongoing ({(current_time - start_time)})"
                    end_datetime_formatted = "Ongoing"
                start_datetime_formatted = start_time.strftime('%m/%d/%Y %H:%M:%S')
                affected_entities_str = sanitize_data_for_excel(", ".join([entity.get("name", "N/A") for entity in problem.get("affectedEntities", [])[:10]]))
                impacted_entities_str = sanitize_data_for_excel(", ".join([entity.get("name", "N/A") for entity in problem.get("impactedEntities", [])[:10]]))
                root_cause_entity = sanitize_data_for_excel(problem.get("rootCauseEntity", {}).get("name", "N/A") if problem.get("rootCauseEntity") else "N/A")
                problem_mgmt_zones = [zone.get("name", "N/A") for zone in problem.get("managementZones", [])]
                for mgmt_zone in problem_mgmt_zones:
                    if mgmt_zone in selected_zones:
                        zone_problem_key = f"{problem_id}_{mgmt_zone}"
                        if zone_problem_key not in zone_processed_problems[mgmt_zone]:
                            zone_processed_problems[mgmt_zone].add(zone_problem_key)
                            problem_record = [
                                sanitize_data_for_excel(problem.get("displayId")),
                                sanitize_data_for_excel(problem_url),
                                sanitize_data_for_excel(problem.get("displayId")),
                                sanitize_data_for_excel(problem.get("title")),
                                sanitize_data_for_excel(problem.get("impactLevel")),
                                sanitize_data_for_excel(problem.get("severityLevel")),
                                sanitize_data_for_excel(problem.get("status")),
                                affected_entities_str,
                                impacted_entities_str,
                                root_cause_entity,
                                sanitize_data_for_excel(mgmt_zone),
                                sanitize_data_for_excel(start_time.strftime('%m/%d/%Y')),
                                sanitize_data_for_excel(start_datetime_formatted),
                                sanitize_data_for_excel(end_datetime_formatted),
                                sanitize_data_for_excel(active_time)
                            ]
                            zone_problems[mgmt_zone].append(problem_record)
            except Exception as e:
                print(f"Error processing problem {problem.get('problemId', 'unknown')}: {e}")
                continue
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        headers = ["Problem ID", "Problem URL", "Problem Link", "Title", "Impact Level", "Severity Level",
                   "Status", "Affected Entities", "Impacted Entities", "Root Cause Entity",
                   "Management Zone", "Date", "Start DateTime", "End DateTime", "Time"]
        for zone_name in selected_zones:
            try:
                sanitized_zone_name = sanitize_sheet_name(zone_name)
                sheet_title = f"Prob_{sanitized_zone_name}"[:31]
                ws = wb.create_sheet(title=sheet_title)
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for problem_data in zone_problems[zone_name]:
                    try:
                        ws.append(problem_data)
                        row_idx = ws.max_row
                        if len(problem_data) > 1 and problem_data[1]:
                            safe_hyperlink(ws, row_idx, 3, problem_data[1], problem_data[2])
                    except Exception as e:
                        print(f"Error adding row to Excel for zone {zone_name}: {e}")
                        continue
            except Exception as e:
                print(f"Error creating sheet for zone '{zone_name}': {e}")
                continue
        try:
            ws_mttr = wb.create_sheet(title="MTTR_Summary")
            mttr_headers = ["Sr No", "Management Zone", "MTTR (Minutes)", "Last 7 Days MTTR (Minutes)", "Total Unique Problems"]
            ws_mttr.append(mttr_headers)
            for cell in ws_mttr[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for mttr_record in mttr_data:
                try:
                    zone_name = mttr_record["Management Zone"]
                    original_zone = None
                    for orig_zone in selected_zones:
                        if sanitize_data_for_excel(orig_zone) == zone_name:
                            original_zone = orig_zone
                            break
                    total_problems = len(zone_problems[original_zone]) if original_zone else 0
                    ws_mttr.append([
                        mttr_record["Sr No"],
                        mttr_record["Management Zone"],
                        mttr_record["MTTR (Minutes)"],
                        mttr_record["Last 7 Days MTTR (Minutes)"],
                        total_problems
                    ])
                except Exception as e:
                    print(f"Error adding MTTR row to Excel: {e}")
                    continue
        except Exception as e:
            print(f"Error creating MTTR sheet: {e}")
        try:
            ws_summary = wb.create_sheet(title="Zone_Summary")
            ws_summary.append(["Management Zone", "Total Unique Problems"])
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            for zone_name in selected_zones:
                ws_summary.append([sanitize_data_for_excel(zone_name), len(zone_problems[zone_name])])
        except Exception as e:
            print(f"Error creating summary sheet: {e}")
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        st.error(f"Failed to generate Excel report: {e}")
        return None

# ===== NEW V3 PROBLEM FETCHER FUNCTIONS (from problem_fetcher_v3.py) =====
def get_week_of_month(date_obj):
    if date_obj is None:
        return "Unknown"
    try:
        first_day = date_obj.replace(day=1)
        dom = date_obj.day
        adjusted_dom = dom + first_day.weekday()
        week = (adjusted_dom - 1) // 7 + 1
        if week == 1:
            return "1st"
        elif week == 2:
            return "2nd"
        elif week == 3:
            return "3rd"
        else:
            return f"{week}th"
    except Exception as e:
        print(f"Error calculating week of month: {e}")
        return "Unknown"

def calculate_mttr_v3(start_time, end_time):
    if end_time is None or end_time == "Ongoing":
        return ">60 min"
    if start_time is None:
        return "Unknown"
    try:
        mttr_minutes = (end_time - start_time).total_seconds() / 60
        if mttr_minutes <= 15:
            return "00-15"
        elif mttr_minutes <= 30:
            return "15-30"
        elif mttr_minutes <= 45:
            return "30-45"
        elif mttr_minutes <= 60:
            return "45-60"
        else:
            return ">60 min"
    except Exception as e:
        print(f"Error calculating MTTR: {e}")
        return "Unknown"

def fetch_problems_parallel_v3(base_url, headers, params, progress_bar, status_text, max_workers=10):
    def fetch_page_with_retry(page_params, max_retries=3):
        for attempt in range(max_retries):
            try:
                response = requests.get(base_url, params=page_params, headers=headers, verify=False, timeout=30)
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 429:
                    wait_time = 2 ** attempt
                    status_text.info(f"Rate limited, waiting {wait_time}s... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                    continue
                else:
                    st.error(f"Error fetching page: {response.status_code} - {response.text}")
                    return None
            except Exception as e:
                st.error(f"Exception fetching page: {e}")
                if attempt == max_retries - 1:
                    return None
                time.sleep(1)
        return None

    try:
        status_text.info("Collecting page keys...")
        first_response = fetch_page_with_retry(params)
        if not first_response:
            raise Exception("Failed to fetch first page")
        
        all_problems = first_response.get("problems", [])
        next_page_key = first_response.get("nextPageKey")
        if not next_page_key:
            return all_problems
        
        page_keys = []
        while next_page_key:
            page_keys.append(next_page_key)
            temp_params = {"nextPageKey": next_page_key}
            temp_response = requests.get(base_url, params=temp_params, headers=headers, verify=False)
            if temp_response.status_code == 200:
                temp_data = temp_response.json()
                next_page_key = temp_data.get("nextPageKey")
                if next_page_key:
                    page_keys.append(next_page_key)
                else:
                    break
            else:
                break
        
        status_text.info(f"Fetching {len(page_keys)} pages in parallel...")
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_page = {
                executor.submit(fetch_page_with_retry, {"nextPageKey": page_key}): page_key
                for page_key in page_keys
            }
            completed = 0
            for future in as_completed(future_to_page):
                result = future.result()
                if result:
                    problems = result.get("problems", [])
                    all_problems.extend(problems)
                completed += 1
                progress_bar.progress(completed / len(page_keys))
        return all_problems
    except Exception as e:
        st.error(f"Error in fetch_problems_parallel_v3: {e}")
        return None

def calculate_mttr_parallel_v3(unique_problems, selected_zones, from_date, to_date):
    def calculate_zone_mttr(zone_name, zone_problems, from_date, to_date):
        current_time = datetime.now()
        total_duration = sum(prob['duration_minutes'] for prob in zone_problems)
        avg_mttr = total_duration / len(zone_problems) if zone_problems else 0
        last_7_days_start = to_date - timedelta(days=7)
        recent_problems = [
            prob for prob in zone_problems
            if prob['start_time'] and prob['start_time'].date() >= last_7_days_start
        ]
        recent_mttr = 0
        if recent_problems:
            recent_duration = sum(prob['duration_minutes'] for prob in recent_problems)
            recent_mttr = recent_duration / len(recent_problems)
        return {
            "Management Zone": zone_name,
            "MTTR (Minutes)": round(avg_mttr, 2),
            "Last 7 Days MTTR (Minutes)": round(recent_mttr, 2)
        }

    zone_problems_dict = defaultdict(list)
    for prob_info in unique_problems.values():
        for zone in prob_info['management_zones']:
            if zone in selected_zones:
                zone_problems_dict[zone].append(prob_info)

    mttr_data = []
    with ThreadPoolExecutor(max_workers=min(8, len(selected_zones))) as executor:
        future_to_zone = {
            executor.submit(calculate_zone_mttr, zone, zone_problems_dict[zone], from_date, to_date): zone
            for zone in selected_zones
        }
        for future in as_completed(future_to_zone):
            result = future.result()
            mttr_data.append(result)
    
    mttr_data.sort(key=lambda x: x["Management Zone"])
    for i, item in enumerate(mttr_data, 1):
        item["Sr No"] = i
    
    return mttr_data

def create_excel_report_v3(all_problems, environment_id, selected_zones, from_date, to_date, status_text):
    try:
        local_timezone = get_localzone()
        current_time = datetime.now()
        
        zone_problems = defaultdict(list)
        zone_processed_problems = defaultdict(set)
        unique_problems = {}

        for problem in all_problems:
            problem_id = problem.get("problemId")
            if not problem_id:
                continue

            try:
                start_time_ms = problem.get("startTime")
                if start_time_ms:
                    start_time = pd.to_datetime(start_time_ms, unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                else:
                    start_time = None
                
                end_time_ms = problem.get("endTime")
                if end_time_ms and end_time_ms != -1:
                    end_time = pd.to_datetime(end_time_ms, unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                else:
                    end_time = None

                if end_time is not None and start_time is not None:
                    duration_minutes = (end_time - start_time).total_seconds() / 60
                elif start_time is not None:
                    duration_minutes = (current_time - start_time).total_seconds() / 60
                else:
                    duration_minutes = 0

                problem_mgmt_zones = [zone.get("name", "N/A") for zone in problem.get("managementZones", [])]
                if not problem_mgmt_zones:
                    problem_mgmt_zones = ["No Management Zone"]
                
                if problem_id not in unique_problems:
                    unique_problems[problem_id] = {
                        'start_time': start_time,
                        'end_time': end_time,
                        'duration_minutes': duration_minutes,
                        'management_zones': problem_mgmt_zones
                    }
                
                problem_url = f"{environment_id}/#problems/problemdetails;pid={problem_id}"

                if end_time is not None and start_time is not None:
                    active_time = str(end_time - start_time)
                    end_datetime_formatted = end_time.strftime('%m/%d/%Y %H:%M:%S')
                elif start_time is not None:
                    active_time = f"Ongoing ({(current_time - start_time)})"
                    end_datetime_formatted = "Ongoing"
                else:
                    active_time = "Unknown"
                    end_datetime_formatted = "Unknown"
                
                start_datetime_formatted = start_time.strftime('%m/%d/%Y %H:%M:%S') if start_time else "Unknown"

                affected_entities_str = sanitize_data_for_excel(", ".join([entity.get("name", "N/A") for entity in problem.get("affectedEntities", [])[:10]]))
                impacted_entities_str = sanitize_data_for_excel(", ".join([entity.get("name", "N/A") for entity in problem.get("impactedEntities", [])[:10]]))
                root_cause_entity = sanitize_data_for_excel(problem.get("rootCauseEntity", {}).get("name", "N/A") if problem.get("rootCauseEntity") else "N/A")
                
                for mgmt_zone in problem_mgmt_zones:
                    if mgmt_zone in selected_zones:
                        zone_problem_key = f"{problem_id}_{mgmt_zone}"
                        if zone_problem_key not in zone_processed_problems[mgmt_zone]:
                            zone_processed_problems[mgmt_zone].add(zone_problem_key)
                            
                            week_of_month = get_week_of_month(start_time)
                            month = start_time.month_name() if start_time else "Unknown"
                            
                            problem_record = [
                                len(zone_problems[mgmt_zone]) + 1,
                                sanitize_data_for_excel(start_time.strftime('%m/%d/%Y')) if start_time else "Unknown",
                                sanitize_data_for_excel(mgmt_zone),
                                sanitize_data_for_excel(problem.get("impactLevel")),
                                sanitize_data_for_excel(problem.get("severityLevel")),
                                sanitize_data_for_excel(start_datetime_formatted),
                                sanitize_data_for_excel(end_datetime_formatted),
                                sanitize_data_for_excel(problem.get("displayId")),
                                sanitize_data_for_excel(problem.get("title")),
                                affected_entities_str,
                                impacted_entities_str,
                                root_cause_entity,
                                sanitize_data_for_excel(problem_url),
                                sanitize_data_for_excel(problem.get("displayId")),
                                sanitize_data_for_excel(problem.get("status")),
                                sanitize_data_for_excel(active_time),
                                sanitize_data_for_excel(calculate_mttr_v3(start_time, end_time)),
                                sanitize_data_for_excel(week_of_month),
                                sanitize_data_for_excel(month)
                            ]
                            zone_problems[mgmt_zone].append(problem_record)

            except Exception as e:
                status_text.error(f"Error processing problem {problem_id}: {e}")
                continue

        mttr_data = calculate_mttr_parallel_v3(unique_problems, selected_zones, from_date, to_date)

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        ws = wb.create_sheet(title="All_Problems")
        headers = [
            "Sl No.", "Date", "Management Zone", "Impact Level", "Severity Level",
            "Start DateTime", "End DateTime", "Problem ID", "Title", "Affected Entities",
            "Impacted Entities", "Root Cause Entity", "Problem URL", "Problem Link",
            "Status", "Time", "MTTR in Min", "Week", "Month"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for zone_name in sorted(zone_problems.keys()):
            if zone_problems[zone_name]:
                ws.append([""] * len(headers))
                separator_row_num = ws.max_row
                ws.cell(row=separator_row_num, column=1, value=f"--- {zone_name} ---").font = Font(bold=True)
                for problem_data in zone_problems[zone_name]:
                    ws.append(problem_data)
                    row_idx = ws.max_row
                    if len(problem_data) > 13 and problem_data[12]:
                        safe_hyperlink(ws, row_idx, 14, problem_data[12], problem_data[13])

        ws_mttr = wb.create_sheet(title="MTTR_Summary")
        mttr_headers = ["Sr No", "Management Zone", "MTTR (Minutes)", "Last 7 Days MTTR (Minutes)", "Total Unique Problems"]
        ws_mttr.append(mttr_headers)
        for cell in ws_mttr[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for mttr_record in mttr_data:
            zone_name = mttr_record["Management Zone"]
            total_problems = len(zone_problems.get(zone_name, []))
            ws_mttr.append([
                mttr_record["Sr No"],
                mttr_record["Management Zone"],
                mttr_record["MTTR (Minutes)"],
                mttr_record["Last 7 Days MTTR (Minutes)"],
                total_problems
            ])

        ws_summary = wb.create_sheet(title="Zone_Summary")
        ws_summary.append(["Management Zone", "Total Unique Problems"])
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for zone_name in sorted(zone_problems.keys()):
            ws_summary.append([sanitize_data_for_excel(zone_name), len(zone_problems[zone_name])])
            
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        st.error(f"Failed to generate Excel report: {e}")
        return None

# ===== GLOBAL ANOMALY DETECTION FUNCTIONS =====
SCHEMA_IDS = [
    "builtin:anomaly-detection.infrastructure-hosts",
    "builtin:anomaly-detection.infrastructure-disks",
    "builtin:anomaly-detection.services",
]

def fetch_environment_level_settings(env_url, api_token, schema_id):
    env_url = env_url.rstrip('/')
    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Content-Type": "application/json; charset=utf-8",
        "Accept": "application/json; charset=utf-8"
    }
    url = (f"{env_url}/api/v2/settings/objects"
           f"?schemaIds={schema_id}&scopes=environment&fields=objectId,value,scope&pageSize=100")
    try:
        response = requests.get(url, headers=headers, verify=False, timeout=30)
        response.raise_for_status()
        return response.json().get("items", [])
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching schema '{schema_id}': {e}")
        return []

def pretty_format_dict(d):
    return json.dumps(d, indent=2)

def render_global_anomaly_tab(env_url, api_token):
    st.header("Global Anomaly Detection Settings")
    st.markdown("Fetch and export environment-level anomaly detection settings from Dynatrace.")
    st.markdown(f"**Using Dynatrace URL:** `{env_url}`")
    st.markdown(f"**Using API Token:** `{'*' * (len(api_token) // 2) if api_token else 'None'}`")
    if st.button("Fetch and Export Anomaly Settings"):
        if not env_url or not api_token:
            st.error("Please provide Dynatrace credentials first.")
            return
        with st.spinner("Fetching anomaly detection settings..."):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Global Anomaly Settings"
                headers = ["Type", "Object ID", "Scope", "Anomaly Detection Details (pretty JSON)"]
                ws.append(headers)
                total_items = 0
                for schema in SCHEMA_IDS:
                    schema_name = schema.split(":")[-1]
                    st.info(f"Fetching settings for: **{schema_name}**")
                    items = fetch_environment_level_settings(env_url, api_token, schema)
                    total_items += len(items)
                    if items:
                        for item in items:
                            obj_id = item.get("objectId", "")
                            scope = item.get("scope", "")
                            value = item.get("value", {})
                            row_data = [schema_name, obj_id, scope, pretty_format_dict(value)]
                            ws.append(row_data)
                        st.success(f"Successfully fetched {len(items)} settings for {schema_name}.")
                    else:
                        st.info("No settings found for {schema_name}.")
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"global_anomaly_settings_{timestamp}.xlsx"
                st.download_button(
                    label="Download Excel Report",
                    data=output.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success(f"Process completed! Exported a total of {total_items} settings.")
            except Exception as e:
                st.error(f"An error occurred during the export process: {e}")

# ===== MAIN APP LOGIC =====
if "auth_role" not in st.session_state:
    st.session_state["auth_role"] = None

if st.session_state.get('rerun_flag'):
    st.session_state['rerun_flag'] = False
    st.rerun()

if st.session_state.get('rerun_after_logout'):
    st.session_state['rerun_after_logout'] = False
    st.rerun()

# Use a container for the animated header
header_container = st.container()

with header_container:
    col1, col2 = st.columns([2, 5])
    with col1:
        st.markdown(f'<div class="logo-animation">', unsafe_allow_html=True)
        try:
            logo = Image.open("apmolens_logo.png")
            st.image(logo, width=280) # Increased logo size to 200
        except Exception:
            pass
    with col2:


        st.markdown(_dedent("""\
<div class="title-wrap">
  <h1 class="title-pro">
    <span class="title-shimmer">Dynatrace Problem Dashboard</span>
  </h1>
  <div class="title-underline"></div>
</div>
"""), unsafe_allow_html=True)


# --- Derived credentials for Admin features ---
env_id = st.session_state.get('auth_dt_url') or st.session_state.get('dt_url_main') or ''
api_token = st.session_state.get('auth_dt_api_token') or st.session_state.get('main_api_token') or ''
# ADMIN PANEL IN THE SIDEBAR
if st.session_state["auth_role"] == "admin":
    # ---- Fancy Admin Panel Header with Avatar & Animation ----
    st.sidebar.markdown(
        """
        <style>
        /* Sidebar look & feel tweaks */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, rgba(20,20,35,0.92), rgba(15,15,25,0.96));
            backdrop-filter: blur(6px);
        }
        section[data-testid="stSidebar"] .admin-card {
            display: flex; align-items: center; gap: 12px;
            padding: 12px 14px; border-radius: 16px;
            background: rgba(255,255,255,0.04);
            box-shadow: 0 2px 14px rgba(0,0,0,0.25) inset, 0 6px 18px rgba(0,0,0,0.16);
            border: 1px solid rgba(255,255,255,0.08);
            cursor: pointer;
            margin-top: 6px;
        }
        .admin-avatar {
            width: 36px; height: 36px; border-radius: 9999px;
            background: radial-gradient(circle at 30% 30%, #8ec5fc, #e0c3fc);
            display: grid; place-items: center;
            box-shadow: 0 4px 10px rgba(0,0,0,0.3), 0 0 0 2px rgba(255,255,255,0.08) inset;
            font-size: 20px; color: #0e0e14; font-weight: 700;
        }
        .admin-text { line-height: 1.1; }
        .admin-title {
            font-size: 1.05rem; font-weight: 800; letter-spacing: .2px;
            background: linear-gradient(90deg, #a5b4fc, #c084fc, #60a5fa, #a5b4fc);
            -webkit-background-clip: text; background-clip: text; color: transparent;
            animation: shimmer 6s linear infinite; background-size: 200% auto;
            text-transform: uppercase;
        }
        .admin-subtitle { font-size: .78rem; opacity: .8; margin-top: 2px; }
        @keyframes shimmer {
            0% { background-position: 0% 50%; }
            100% { background-position: 200% 50%; }
        }
        .admin-features {
            max-height: 0; overflow: hidden; opacity: 0;
            transform: translateY(-6px);
            transition: max-height .45s ease, opacity .35s ease, transform .35s ease;
            margin: 6px 2px 0 2px;
        }
        .admin-card:hover + .admin-features,
        .admin-features:hover {
            max-height: 900px; opacity: 1; transform: translateY(0);
        }
        section[data-testid="stSidebar"] [data-testid="stRadio"] {
            padding: 10px 12px; border-radius: 14px;
            background: rgba(255,255,255,0.04);
            border: 1px solid rgba(255,255,255,0.08);
            box-shadow: 0 8px 26px rgba(0,0,0,0.24);
        }
        section[data-testid="stSidebar"] [data-testid="stRadio"] label {
            padding: 6px 8px; border-radius: 10px;
        }
        section[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
            background: rgba(255,255,255,0.06);
        }
        </style>
        <div class="admin-card">
            <div class="admin-avatar">U</div>
            <div class="admin-text">
                <div class="admin-title">Admin Panel</div>
                <div class="admin-subtitle">Manage & configure</div>
            </div>
        </div>
        """, unsafe_allow_html=True
    )
    if "admin_tab" not in st.session_state:
        st.session_state["admin_tab"] = "Dashboard"
    
    admin_options = ["Dashboard", "Add User", "Manage Users", "OneAgent Deployment", "Global Anomaly Detection", "Maintenance Window"]
    st.sidebar.markdown('<div class="admin-features">', unsafe_allow_html=True)
    selected_admin_option = st.sidebar.radio(
        "Select an Admin Feature:",
        options=admin_options,
        index=admin_options.index(st.session_state["admin_tab"]),
        key="admin_radio"
    )
    st.sidebar.markdown("</div>", unsafe_allow_html=True)
    st.session_state["admin_tab"] = selected_admin_option
    
    st.sidebar.markdown("---")
    st.sidebar.button("Logout", on_click=logout, use_container_width=True)

    if selected_admin_option == "Add User":
        render_add_user_tab()
        
    elif selected_admin_option == "Manage Users":
        admin_panel_tab()
    elif selected_admin_option == "OneAgent Deployment":
        render_oneagent_deployment_tab()
    elif selected_admin_option == "Global Anomaly Detection":

        render_global_anomaly_tab(env_id, api_token)

    elif selected_admin_option == "Maintenance Window":
        maintenance_window_ui(env_id, api_token)


    # If admin is not on Dashboard, stop rendering the rest of the dashboard
    if st.session_state.get("auth_role") == "admin" and st.session_state.get("admin_tab") != "Dashboard":
        st.stop()


show_dashboard = (st.session_state.get("auth_role") == "user") or (st.session_state.get("auth_role") == "admin" and st.session_state.get("admin_tab") == "Dashboard")
if show_dashboard:
    # --- Sticky Action Bar ---
    st.markdown("<div class='sticky-bar'></div>", unsafe_allow_html=True)
    col_sb1, col_sb2, col_sb3 = st.columns([1,1,1])
    with col_sb1:
        if st.button("🔄 Fetch Problems", use_container_width=True):
            st.session_state['trigger_fetch'] = True
    with col_sb2:
        if st.button("⬇️ Export Excel", use_container_width=True):
            st.session_state['trigger_export'] = True
    with col_sb3:
        if st.button("🧹 Reset Filters", use_container_width=True):
            for k in list(st.session_state.keys()):
                if k.endswith("_filter"):
                    st.session_state.pop(k)
            st.rerun()

if show_dashboard:
    # --- KPI Cards (quick overview) ---
    import time as _time
    def _kpi_row(metrics):
        c1,c2,c3,c4 = st.columns(4)
        for (title, value, trend), col in zip(metrics, (c1,c2,c3,c4)):
            with col:
                st.markdown(f"""
                <div class="kpi">
                  <div class="kpi-title">{title}</div>
                  <div class="kpi-value">{value}</div>
                  <div class="kpi-trend">{trend}</div>
                </div>
                """, unsafe_allow_html=True)

    _kpi_placeholder = st.empty()
    with _kpi_placeholder.container():
        s1,s2,s3,s4 = st.columns(4)
        for s in (s1,s2,s3,s4):
            with s: st.markdown('<div class="skel"></div>', unsafe_allow_html=True)

_time.sleep(0.05)
_df  = st.session_state.get("df", pd.DataFrame())
_fdf = st.session_state.get("filtered_df", pd.DataFrame())

# Totals
total_loaded   = int(_df.get("Problem ID", pd.Series(dtype=str)).nunique()) if not _df.empty else 0
total_filtered = int(_fdf.get("Problem ID", pd.Series(dtype=str)).nunique()) if not _fdf.empty else 0

open_problems  = int((_fdf["Status"] == "OPEN").sum()) if ("Status" in _fdf) else 0
zones_loaded   = len(st.session_state.get("management_zones", []))
mttr_avg       = st.session_state.get("avg_mttr_min", 0)

_kpi_placeholder.empty()
_kpi_row([
  ("Total Problems (filtered)", f"{total_filtered:,}", "after filters"),
  ("Open Problems", f"{open_problems:,}", "needs attention"),
  ("Zones Loaded", f"{zones_loaded}", "from API/CSV"),
  ("Avg MTTR (min)", f"{mttr_avg}", "computed")
])


# Main App Logic starts here
if show_dashboard:
    st.subheader("Dynatrace API Credentials")
    c1, c2 = st.columns([3, 2])
    is_admin = st.session_state.get("auth_role") == "admin"
    with c1:
        default_env = st.session_state.get('auth_dt_url', '')
        env_id = st.text_input(
            "Dynatrace Environment URL",
            value=default_env,
            disabled=(not is_admin),
            key="main_env_id",
            help="Example: https://your-tenant.live.dynatrace.com"
        )
    with c2:
        api_token = st.session_state.get('auth_dt_api_token', '')
        api_token_input = st.text_input(
            "API Token",
            value=api_token if is_admin else "********",
            type="password",
            disabled=(not is_admin),
            key="main_api_token",
            help="API Token with 'Installer download' permission."
        )

    if is_admin and env_id and api_token_input:
        st.session_state['auth_dt_url'] = env_id
        st.session_state['auth_dt_api_token'] = api_token_input
    elif not is_admin:
        pass
    else:
        st.warning("Admin: Please enter your Dynatrace credentials.")

    # USER & ADMIN DASHBOARD LOGIC
    if st.session_state.get("auth_role") == "user" or (st.session_state.get("auth_role") == "admin" and st.session_state.get("admin_tab") == "Dashboard"):
        if st.session_state.get("auth_role") == "user":
            st.sidebar.button("Logout", on_click=logout, use_container_width=True)
    
        st.markdown("For SaaS: https://abc12345.live.dynatrace.com/")
        management_zones = st.session_state.get('management_zones', [])
        if st.button("Fetch Management Zones"):
            if env_id and api_token:
                mgmt_zones = fetch_management_zones(env_id, api_token)
                st.session_state['management_zones'] = mgmt_zones
                st.session_state['all_management_zones'] = mgmt_zones
                management_zones = mgmt_zones
                st.success(f"Fetched {len(mgmt_zones)} management zones.")
                st.session_state['zones_loaded_at'] = _time.time()
                st.rerun()
            else:
                st.error("Enter valid Dynatrace environment and API token.")
        all_management_zones = st.session_state.get('management_zones', [])
        st.subheader("Set Problem Fetch Parameters")
        use_all_mz = st.checkbox("No Management Zone (Fetch All)", value=(st.session_state.get("auth_role") == "admin"))
        mz_options = (all_management_zones if st.session_state.get("auth_role") == "admin" else st.session_state.get("auth_zones", []))
        selected_zones = st.multiselect("Select Management Zones", options=mz_options, disabled=use_all_mz)
        c3, c4 = st.columns(2)
        with c3:
            date_from = st.date_input("From Date", value=datetime.today() - timedelta(days=14), key="from_main")
            hour_from = st.selectbox("From Hour", options=list(range(24)), format_func=lambda x: f"{x:02d}", key="hfrom_main")
            minute_from = st.selectbox("From Minute", options=list(range(0, 60, 5)), format_func=lambda x: f"{x:02d}", key="mfrom_main")
        with c4:
            date_to = st.date_input("To Date", value=datetime.today(), key="to_main")
            hour_to = st.selectbox("To Hour", options=list(range(24)), format_func=lambda x: f"{x:02d}", index=23, key="hto_main")
            minute_to = st.selectbox("To Minute", options=list(range(0, 60, 5)), format_func=lambda x: f"{x:02d}", index=11, key="mto_main")

        def fetch_problems(env_id, api_token, mz_list, use_all_mz, date_from, hour_from, minute_from, date_to, hour_to, minute_to):
            headers = {"Authorization": f"Api-Token {api_token}"}
            local_timezone = datetime.now().astimezone().tzinfo
            def convert_to_timestamp(date, hour, minute):
                dt_obj = datetime.combine(date, datetime.min.time()) + timedelta(hours=hour, minutes=minute)
                local_dt = pd.Timestamp(dt_obj).tz_localize(local_timezone)
                dt_utc = local_dt.tz_convert('UTC')
                return int(dt_utc.timestamp() * 1000)
            from_ts = convert_to_timestamp(date_from, hour_from, minute_from)
            to_ts = convert_to_timestamp(date_to, hour_to, minute_to)
            params = {"from": from_ts, "to": to_ts, "pageSize": 10, "fields": "+managementZones"}
            if not use_all_mz and mz_list:
                mz_str = ", ".join([f'"{mz}"' for mz in mz_list])
                params["problemSelector"] = f'managementZones({mz_str})'
            base_url = f"{env_id.rstrip('/')}/api/v2/problems"
            all_problems = []
            with st.spinner("Fetching Problems..."):
                while True:
                    resp = requests.get(base_url, params=params, headers=headers, verify=False)
                    if resp.status_code != 200:
                        st.error(f"Error: {resp.status_code}\n{resp.text}")
                        break
                    data = resp.json()
                    problems = data.get("problems", [])
                    all_problems.extend(problems)
                    next_page_key = data.get("nextPageKey")
                    if not next_page_key:
                        break
                    params = {"nextPageKey": next_page_key}
            problem_rows = []
            for problem in all_problems:
                start_time = pd.to_datetime(problem.get("startTime"), unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                end_time_raw = problem.get("endTime")
                if end_time_raw and end_time_raw != -1:
                    end_time = pd.to_datetime(end_time_raw, unit="ms").tz_localize("UTC").tz_convert(local_timezone).replace(tzinfo=None)
                else:
                    end_time = "Ongoing"
                if end_time == "Ongoing":
                    mttr = ">60"
                else:
                    mttr_min = (end_time - start_time).total_seconds() / 60
                    if mttr_min <= 15:
                        mttr = "00-15"
                    elif mttr_min <= 30:
                        mttr = "15-30"
                    elif mttr_min <= 45:
                        mttr = "30-45"
                    elif mttr_min <= 60:
                        mttr = "45-60"
                    else:
                        mttr = ">60"
                problem_url = f"{env_id.rstrip('/')}/#problems/problemdetails;pid={problem.get('problemId')}"
                management_zones = ", ".join([z.get("name", "N/A") for z in problem.get("managementZones", [])])
                month = start_time.strftime("%b")
                affected_entities = [e.get("name", "N/A") for e in problem.get("affectedEntities", [])] or ["N/A"]
                impacted_entities = [e.get("name", "N/A") for e in problem.get("impactedEntities", [])] or ["N/A"]
                root_cause_entity = problem.get("rootCauseEntity", {}).get("name", "N/A") if problem.get("rootCauseEntity") else "N/A"
                for affected in affected_entities:
                    for impacted in impacted_entities:
                        problem_rows.append([
                            problem.get("displayId"),
                            problem_url,
                            problem.get("displayId"),
                            problem.get("title"),
                            problem.get("impactLevel"),
                            problem.get("severityLevel"),
                            problem.get("status"),
                            affected,
                            impacted,
                            root_cause_entity,
                            management_zones,
                            start_time.strftime('%d-%m-%Y'),
                            start_time.strftime('%m-%d-%Y %H:%M:%S'),
                            end_time if end_time == "Ongoing" else end_time.strftime('%m-%d-%Y %H:%M:%S'),
                            (start_time.day - 1) // 7 + 1,
                            end_time if end_time == "Ongoing" else (end_time - start_time),
                            mttr,
                            month
                        ])
            cols = [
                "Problem ID", "Problem URL", "Problem Link", "Title", "Impact Level", "Severity Level",
                "Status", "Affected Entities", "Impacted Entities", "Root Cause Entity",
                "Management Zones", "Date", "Start DateTime", "End DateTime", "Week", "Time", "MTTR in Min", "Month"
            ]
            df = pd.DataFrame(problem_rows, columns=cols)
            if st.session_state.get("auth_role") != "admin":
                allowed_zones = st.session_state.get("auth_zones", [])
                df = df[df["Management Zones"].apply(lambda s: bool(set([z.strip() for z in s.split(',')]) & set(allowed_zones)))]
            return df

    # ===== PROBLEM TABLE, FILTERS, CHARTS, HEATMAPS & PDF EXPORT =====
    st.sidebar.header("Upload / Download Data")
    uploaded_file = st.sidebar.file_uploader("Upload exported Problems CSV", type=["csv"])
    if "df" not in st.session_state:
        st.session_state["df"] = pd.DataFrame()
    df = st.session_state["df"]
    if uploaded_file:
        df = pd.read_csv(uploaded_file)
        st.session_state["df"] = df
        st.session_state["filtered_df"] = df.copy()
        st.session_state["last_upload"] = _time.time()
        # removed explicit st.rerun(); Streamlit reruns automatically on widget change.
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors='coerce', dayfirst=True)
    else:
        if df.empty:
            st.warning("Please upload a Problems CSV file exported from the app or fetch data via API fetcher below.")
    if not df.empty:
        st.sidebar.header("🔍 Filters")
        filters = {}
        filter_columns = ["Week", "Month", "Severity Level", "Impact Level", "Affected Entities", "Management Zones", "Status", "MTTR in Min", "Title"]
        for column in filter_columns:
            col_filtered = df[df["Management Zones"].isin(st.session_state.get("auth_zones", []))] if column == "Management Zones" and st.session_state.get("auth_role") != "admin" else df
            options = sorted(col_filtered[column].dropna().astype(str).unique())
            selected = st.sidebar.multiselect(f"{column}", options=options, key=column+"_filter")
            filters[column] = selected
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        f_min_dt = df["Date"].min()
        f_max_dt = df["Date"].max()
        date_range = st.sidebar.date_input(
            "Date range filter:",
            value=(f_min_dt, f_max_dt) if pd.notnull(f_min_dt) and pd.notnull(f_max_dt) else (datetime.today(), datetime.today()),
            key='date_filter'
        )
        if st.sidebar.button("Apply Filters", key="apply_filters_btn"):
            mask = pd.Series([True]*len(df))
            if isinstance(date_range, (tuple, list)):
                start, end = date_range
            else:
                start = end = date_range
            mask &= (df["Date"] >= pd.to_datetime(start)) & (df["Date"] <= pd.to_datetime(end))
            for col, vals in filters.items():
                if vals:
                    mask &= df[col].astype(str).isin(vals)
            filtered_df = df[mask]
            st.session_state['filtered_df'] = filtered_df.copy()
            st.session_state['last_filters_apply'] = _time.time()
            st.rerun()
    else:
        st.session_state['filtered_df'] = pd.DataFrame()
    if st.button("Fetch Problems", key="fetch_main"):
        if env_id and api_token:
            api_zones = [] if use_all_mz else selected_zones
            df = fetch_problems(
                env_id, api_token, api_zones, use_all_mz,
                date_from, hour_from, minute_from, date_to, hour_to, minute_to
            )
            st.session_state["df"] = df
            st.session_state['filtered_df'] = df
            st.success("Problems fetched and loaded! Filters reset.")
            st.rerun()
        else:
            st.error("Dynatrace environment URL or API token not available. Please contact your administrator.")
    df = st.session_state.get("df", pd.DataFrame())
    filtered_df = st.session_state.get("filtered_df", df)
    st.header("Problem Fetcher and Dashboard")
    st.write("### Loaded Dataset (Filtered view)")
# --- KPIs ---
try:
    _mttr_series = None
    if "Duration Minutes" in filtered_df.columns:
        _mttr_series = pd.to_numeric(filtered_df["Duration Minutes"], errors="coerce")
    elif "Time" in filtered_df.columns and pd.api.types.is_timedelta64_dtype(filtered_df["Time"]):
        _mttr_series = filtered_df["Time"].dt.total_seconds() / 60
    if _mttr_series is not None:
        _mttr = _mttr_series.dropna()
        avg_mttr = round(_mttr.mean(), 2) if not _mttr.empty else None
        med_mttr = round(_mttr.median(), 2) if not _mttr.empty else None
        cnt_resolved = int((_mttr_series.notna()).sum()) if _mttr_series is not None else 0
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("MTTR (avg, minutes)", value=avg_mttr if avg_mttr is not None else "—")
        with c2:
            st.metric("MTTR (median, minutes)", value=med_mttr if med_mttr is not None else "—")
        with c3:
            st.metric("Resolved problems counted", value=cnt_resolved)
except Exception as _e:
    st.caption(f"MTTR metrics unavailable: {_e}")
st.dataframe(_make_arrow_safe(filtered_df), use_container_width=True)

csv = filtered_df.to_csv(index=False)
st.download_button(label="⬇️ Download CSV", data=csv, file_name="problems.csv", mime='text/csv')

def shorten_label(label, max_len=16):
    label = str(label)
    return label if len(label) <= max_len else label[:max_len-2] + '..'

def download_pdf_report(charts, heatmaps, summary_text=None):
    from fpdf import FPDF
    import tempfile
    from PIL import Image
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=10)
    page_width = 210
    page_height = 297
    margin = 10
    padding = 7
    max_img_width = page_width - 2 * margin
    images_per_page = 4
    max_img_height = (page_height - 2 * margin - padding * (images_per_page - 1)) / images_per_page
    images = charts + heatmaps
    if summary_text:
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 7, summary_text)
    for i in range(0, len(images), images_per_page):
        pdf.add_page()
        page_images = images[i:i + images_per_page]
        y_position = margin
        for img_bytes in page_images:
            with tempfile.NamedTemporaryFile(delete=True, suffix=".png") as tmp_img:
                tmp_img.write(img_bytes)
                tmp_img.flush()
                img = Image.open(tmp_img.name)
                iw, ih = img.size
                dpi = img.info.get('dpi', (96, 96))[0]
                iw_mm = (iw / dpi) * 25.4
                ih_mm = (ih / dpi) * 25.4
                scale = min(max_img_width / iw_mm, max_img_height / ih_mm, 1)
                disp_w = iw_mm * scale
                disp_h = ih_mm * scale
                x_position = (page_width - disp_w) / 2
                pdf.image(tmp_img.name, x=x_position, y=y_position, w=disp_w, h=disp_h)
                y_position += disp_h + padding
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    return pdf_bytes

# Main tabs for Charts, Heatmap, Maintenance Window, Split by MZ, and V3 Fetcher
tab_charts, tab_heatmaps, tab_split_by_mz, tab_problem_fetcher_v3 = st.tabs(["Charts", "Heatmap & Insights", "Split by MZ", "Problem Fetcher v3.0"])

with tab_charts:
    chart_df = st.session_state['filtered_df'] if not st.session_state['filtered_df'].empty else pd.DataFrame()
    bar_chart_imgs = []
    if not chart_df.empty:
        st.header("Charts (based on Applied Filters)")
        dfc = chart_df.copy()
        if "Date" in dfc.columns:
            dfc["Date"] = pd.to_datetime(dfc["Date"], errors="coerce")
            dfc = dfc.dropna(subset=["Date"])
            date_group = dfc.groupby(dfc["Date"].dt.date)["Problem ID"].nunique().sort_index()
            avg_val = date_group.mean()
            fig, ax = plt.subplots(figsize=(13, 4))
            date_labels = [f"({d.strftime('%a')[0]}) {d.strftime('%d-%b')}" for d in date_group.index]
            bars = ax.bar(date_labels, date_group.values, width=0.8)
            for p in bars:
                ax.annotate(str(int(p.get_height())), (p.get_x() + p.get_width() / 2, p.get_height()), ha='center', va='bottom', fontsize=8, fontweight='bold')
            ax.axhline(avg_val, color='red', linestyle='--', linewidth=2, label=f'Avg: {avg_val:.1f}')
            ax.legend()
            ax.set_title("Date-wise Problem Distribution")
            ax.set_xlabel("Date")
            ax.set_ylabel("Unique Problem IDs")
            ax.tick_params(axis='x', rotation=90, labelsize=8)
            plt.subplots_adjust(bottom=0.25)
            st.pyplot(fig)
            buf = io.BytesIO()
            fig.savefig(buf, format="png")
            buf.seek(0)
            bar_chart_imgs.append(buf.getvalue())
            plt.close(fig)
        chart_columns = ["Severity Level", "Impact Level", "Title", "Affected Entities", "Impacted Entities", "MTTR in Min"]
        for col in chart_columns:
            if col not in dfc.columns:
                continue
            vc = dfc.groupby(col)["Problem ID"].nunique().sort_values(ascending=False)
            if col in ["Title", "Affected Entities", "Impacted Entities"]:
                vc = vc.head(10)
            fig, ax = plt.subplots(figsize=(13, 4))
            bars = ax.bar(vc.index.astype(str), vc.values, width=0.8)
            for p in bars:
                ax.annotate(str(int(p.get_height())), (p.get_x() + p.get_width() / 2, p.get_height()), ha='center', va='bottom', fontsize=8, fontweight='bold')
            ax.set_title(f"{col} Distribution")
            ax.set_xlabel(col)
            ax.set_ylabel("Unique Problem IDs")
            ax.tick_params(axis='x', rotation=0, labelsize=8)
            labels = [label.get_text() for label in ax.get_xticklabels()]
            max_len = 12
            new_labels = ['\n'.join([text[i:i + max_len] for i in range(0, len(text), max_len)]) for text in labels]
            ax.set_xticks(range(len(new_labels)))
            ax.set_xticklabels(new_labels)
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.25)
            st.pyplot(fig)
            buf = io.BytesIO()
            fig.savefig(buf, format="png")
            buf.seek(0)
            bar_chart_imgs.append(buf.getvalue())
            plt.close(fig)
    else:
        st.info("Load and apply filters to display charts.")
    st.session_state['bar_chart_imgs'] = bar_chart_imgs
with tab_heatmaps:
    heatmap_df = filtered_df
    heatmap_imgs = []
    summary_lines = []
    st.header(" Heatmap & Insights (Filtered Results or Last 14 Days)")
    if not heatmap_df.empty:
        st.subheader("Top 10 Titles vs Date")
        heatmap_df["Start DateTime"] = pd.to_datetime(heatmap_df["Start DateTime"], errors="coerce")
        recent_df = heatmap_df.copy()
        max_date = recent_df["Start DateTime"].max()
        if pd.notnull(max_date):
            min_date = max_date - timedelta(days=13)
            recent_df = recent_df[recent_df["Start DateTime"].dt.date >= min_date.date()]
        top_titles = recent_df["Title"].value_counts().head(10).index
        title_date_count = recent_df[recent_df["Title"].isin(top_titles)].groupby(
            [recent_df["Start DateTime"].dt.date, "Title"]).size().unstack(fill_value=0)
        if not title_date_count.empty:
            fig1, ax1 = plt.subplots(figsize=(11, 4.6))
            cmap = sns.diverging_palette(145, 10, s=90, l=40, as_cmap=True)
            sns.heatmap(
                title_date_count,
                annot=True,
                fmt="d",
                cmap=cmap,
                ax=ax1,
                cbar_kws={'label': "Problem Count"},
                annot_kws={'size': 9, 'weight': 'bold'}
            )
            ax1.set_title("Heatmap: Top 10 Titles by Date", fontsize=12, fontweight='bold')
            ax1.set_xlabel("Problem Title", fontsize=10)
            ax1.set_ylabel("Date", fontsize=10)
            def shorten_label(label, length=15):
                return label if len(label) <= length else label[:length-3] + "..."
            ax1.set_xticklabels([shorten_label(x.get_text(), 15) for x in ax1.get_xticklabels()], rotation=30, fontsize=8)
            fig1.tight_layout()
            st.pyplot(fig1)
            buf1 = io.BytesIO()
            fig1.savefig(buf1, format='png', bbox_inches='tight', dpi=140)
            heatmap_imgs.append(buf1.getvalue())
            plt.close(fig1)
        else:
            st.info("No title/date combinations for heatmap.")
        st.subheader("Problems by Date and Hour")
        recent_df["Hour"] = recent_df["Start DateTime"].dt.hour
        recent_df["Day"] = recent_df["Start DateTime"].dt.strftime('%d-%b')
        hour_pivot = recent_df.pivot_table(index="Day", columns="Hour", values="Problem ID", aggfunc="nunique", fill_value=0)
        if not hour_pivot.empty:
            fig2, ax2 = plt.subplots(figsize=(13, 5))
            sns.heatmap(hour_pivot, annot=True, cmap="YlOrRd", fmt="d", linewidths=0.5, ax=ax2, cbar_kws={'label': "Unique Problem Count"}, annot_kws={'size': 9, 'weight': 'bold'})
            ax2.set_title("Heatmap: Problem Count by Date and Hour", fontsize=12, fontweight='bold')
            ax2.set_xlabel("Hour", fontsize=10)
            ax2.set_ylabel("Date", fontsize=10)
            fig2.tight_layout()
            st.pyplot(fig2)
            buf2 = io.BytesIO()
            fig2.savefig(buf2, format="png", bbox_inches='tight', dpi=140)
            heatmap_imgs.append(buf2.getvalue())
            plt.close(fig2)
            max_val = hour_pivot.max().max()
            max_locs = [(day, hour) for day in hour_pivot.index for hour in hour_pivot.columns if hour_pivot.loc[day, hour] == max_val]
            summary_lines.append(f"Max problems ({int(max_val)}) occurred at:")
            for d, h in max_locs:
                summary_lines.append(f"Date: {d} Hour: {h:02d}")
            peak_hours = hour_pivot.sum().nlargest(3)
            summary_lines.append("Top 3 Peak hours:")
            for hour, count in peak_hours.items():
                summary_lines.append(f"Hour {hour:02d} with {int(count)} problems")
            peak_days = hour_pivot.sum(axis=1).nlargest(3)
            summary_lines.append("Top 3 Peak days:")
            for day, count in peak_days.items():
                summary_lines.append(f"{day} with {int(count)} problems")
        else:
            st.info("Not enough data for hourly heatmap.")
    else:
        st.info("Apply filters or fetch data to visualize heatmaps.")
    st.subheader("Overall Heatmap Analysis")
    if summary_lines:
        overall_analysis_md = """
**Heatmap Analysis Summary:**
Below are key insights derived from the heatmap visualization:
""" + "\n".join(f"- {line}" for line in summary_lines)
    else:
        overall_analysis_md = "**Heatmap Analysis Summary:**\n\nNo significant insights available from the current data/filter set."
    st.markdown(f"""
    <div style='padding: 10px; background-color: #21232b; border-radius: 13px; margin-bottom:10px; color: #f8fcf7; font-size: 1rem'>
    {overall_analysis_md}
    </div>
    """, unsafe_allow_html=True)
    if st.button("Download Analysis PDF Report"):
        export_df = filtered_df.copy()
        pdf_data = download_pdf_report(st.session_state.get('bar_chart_imgs', []), heatmap_imgs, overall_analysis_md)
        st.download_button("Download PDF Report", data=pdf_data, file_name="Dynatrace_Problems_Report.pdf", mime='application/pdf')
# with tab_maintenance_window:
#     maintenance_window_ui(env_id, api_token)
with tab_split_by_mz:
    st.header("Dynatrace Data Fetcher - Unique Problems per Management Zone")
    dynatrace_url = st.session_state.get('auth_dt_url', '')
    api_token = st.session_state.get('auth_dt_api_token', '')
    if not dynatrace_url or not api_token:
        st.warning("Please enter your Dynatrace credentials in the main dashboard to use this feature.")
        st.stop()
    if st.button("Get Management Zones", key="smz_get_mz_button"):
        zones = fetch_management_zones(dynatrace_url, api_token)
        st.session_state['smz_zones'] = zones
        if zones:
            st.success(f"Fetched {len(zones)} management zones.")
            st.session_state['zones_loaded_at'] = _time.time()
            st.rerun()
        else:
            st.warning("No management zones found or an error occurred.")
    selected_zones = []
    if 'smz_zones' in st.session_state and st.session_state.smz_zones:
        selected_zones = st.multiselect("Select Management Zones", st.session_state.smz_zones, key="smz_selected_zones")
    elif 'smz_zones' in st.session_state and not st.session_state.smz_zones:
        st.info("No management zones available. Please click 'Get Management Zones' to fetch them first.")
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", key="smz_from_date", value=datetime.today() - timedelta(days=7))
        from_time = st.time_input("From Time", value=time(0, 0), key="smz_from_time")
    with col2:
        to_date = st.date_input("To Date", key="smz_to_date", value=datetime.today())
        to_time = st.time_input("To Time", value=time(23, 59), key="smz_to_time")
    if st.button("Fetch Data and Create Excel", key="smz_fetch_data_button"):
        if not selected_zones:
            st.error("Please select at least one management zone.")
        else:
            with st.spinner("Processing data and generating Excel..."):
                excel_data = process_data_and_create_excel(
                    dynatrace_url, api_token, selected_zones,
                    from_date, from_time, to_date, to_time
                )
            if excel_data:
                st.download_button(
                    label="Download Excel Report",
                    data=excel_data,
                    file_name="Dynatrace_Problems_By_MZ.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
with tab_problem_fetcher_v3:
    st.header("Dynatrace Problem Fetcher v3.0 - Single Consolidated Excel")
    dynatrace_url = st.session_state.get('auth_dt_url', '')
    api_token = st.session_state.get('auth_dt_api_token', '')
    if not dynatrace_url or not api_token:
        st.warning("Please enter your Dynatrace credentials in the main dashboard to use this feature.")
        st.stop()
    
    st.markdown("This tool fetches problems and generates a single Excel file with a main problem sheet, an MTTR summary, and a zone summary.")

    if st.button("Get Management Zones", key="v3_get_mz_button"):
        with st.spinner("Fetching management zones..."):
            zones = fetch_management_zones(dynatrace_url, api_token)
            st.session_state['v3_zones'] = zones
        if zones:
            st.success(f"Fetched {len(zones)} management zones.")
            st.session_state['zones_loaded_at'] = _time.time()
            st.rerun()
        else:
            st.warning("No management zones found or an error occurred.")

    use_all_mz_v3 = st.checkbox("No Management Zone (Fetch All Problems)", key="v3_no_mz")
    selected_zones_v3 = []
    if 'v3_zones' in st.session_state and st.session_state.v3_zones:
        selected_zones_v3 = st.multiselect("Select Management Zones", sorted(st.session_state.v3_zones), disabled=use_all_mz_v3, key="v3_selected_zones")
    elif not use_all_mz_v3:
        st.info("No management zones available. Please click 'Get Management Zones' to fetch them first.")
    
    col1, col2 = st.columns(2)
    with col1:
        from_date_v3 = st.date_input("From Date", key="v3_from_date", value=datetime.today() - timedelta(days=7))
        from_time_v3 = st.time_input("From Time", value=time(0, 0), key="v3_from_time")
    with col2:
        to_date_v3 = st.date_input("To Date", key="v3_to_date", value=datetime.today())
        to_time_v3 = st.time_input("To Time", value=time(23, 59), key="v3_to_time")

    if st.button("Fetch Data and Create Excel Report", key="v3_fetch_data_button"):
        if not (use_all_mz_v3 or selected_zones_v3):
            st.error("Please select at least one management zone or check the 'No Management Zone' option.")
        else:
            if not use_all_mz_v3 and "No Management Zone" in selected_zones_v3:
                st.warning("The 'No Management Zone' option is not compatible with selecting specific zones. Please use one or the other.")
            
            with st.spinner("Processing data and generating Excel..."):
                from_datetime = datetime.combine(from_date_v3, from_time_v3)
                to_datetime = datetime.combine(to_date_v3, to_time_v3)
                
                mz_to_fetch = selected_zones_v3 if not use_all_mz_v3 else st.session_state.get('v3_zones', [])
                
                progress_bar = st.progress(0, text="Starting data fetch...")
                status_text = st.empty()

                if not mz_to_fetch and not use_all_mz_v3:
                     st.error("Please select management zones or choose to fetch all.")
                else:
                    try:
                        # FIX: Rewriting the problematic f-string expression
                        if not use_all_mz_v3:
                            zone_list_str = [f'"{zone}"' for zone in mz_to_fetch]
                            problem_selector = f'managementZones({", ".join(zone_list_str)})'
                        else:
                            problem_selector = None
                        
                        def convert_to_timestamp_v3(date, time_obj):
                            dt_obj = datetime.combine(date, time_obj)
                            local_dt = pd.Timestamp(dt_obj).tz_localize(get_localzone())
                            utc_dt = local_dt.tz_convert('UTC')
                            return int(utc_dt.timestamp() * 1000)

                        params = {
                            "from": convert_to_timestamp_v3(from_date_v3, from_time_v3),
                            "to": convert_to_timestamp_v3(to_date_v3, to_time_v3),
                            "pageSize": 10,
                            "fields": "+managementZones"
                        }
                        if problem_selector:
                            params["problemSelector"] = problem_selector
                        
                        all_problems = fetch_problems_parallel_v3(
                            base_url=f'{dynatrace_url.rstrip("/")}/api/v2/problems',
                            headers={"Authorization": f"Api-Token {api_token}"},
                            params=params,
                            progress_bar=progress_bar,
                            status_text=status_text,
                            max_workers=10
                        )

                        if all_problems:
                            status_text.success(f"Fetched {len(all_problems)} problems. Generating report...")
                            # Create the Excel report with MTTR and Summary sheets
                            excel_data = create_excel_report_v3(
                                all_problems,
                                dynatrace_url,
                                selected_zones_v3 if not use_all_mz_v3 else st.session_state.get('v3_zones', []),
                                from_date_v3, to_date_v3,
                                status_text
                            )
                            if excel_data:
                                st.download_button(
                                    label="Download Consolidated Excel Report",
                                    data=excel_data,
                                    file_name="Dynatrace_Problems_v3.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                status_text.success("Report generated successfully.")
                            else:
                                status_text.error("Failed to generate the Excel report.")
                        else:
                            status_text.warning("No problems found for the selected criteria.")
                    except Exception as e:
                        st.error(f"An unexpected error occurred: {e}")
                    finally:
                        progress_bar.empty()
                        gc.collect()


st.markdown("""
<style>
/* Hide fullscreen/zoom buttons globally */
button[title="View fullscreen"],
button[aria-label="View fullscreen"],
div[title="View fullscreen"],
[data-testid="StyledFullScreenButton"] { display: none !important; }
</style>
""", unsafe_allow_html=True)